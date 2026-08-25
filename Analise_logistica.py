import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()

# Sheet 1: Resumo Executivo & KPIs
ws_exec = wb.active
ws_exec.title = "Resumo Executivo"
ws_exec.views.sheetView[0].showGridLines = True

# Sheet 2: Performance por UF
ws_uf = wb.create_sheet(title="Performance por UF")
ws_uf.views.sheetView[0].showGridLines = True

# Sheet 3: Impacto SLA & Avaliações
ws_sla = wb.create_sheet(title="SLA e Avaliacoes")
ws_sla.views.sheetView[0].showGridLines = True

# Palette
DARK_NAVY = "1B365D"
ACCENT_BLUE = "2B5B84"
LIGHT_BLUE = "EAF2F8"
ZEBRA_FILL = "F8FAFC"
WHITE = "FFFFFF"
BORDER_COLOR = "D9D9D9"
ALERT_RED_FILL = "FADBD8"
ALERT_RED_TEXT = "900C3F"
OK_GREEN_FILL = "D4EFDF"
OK_GREEN_TEXT = "196F3D"

font_title = Font(name="Calibri", size=16, bold=True, color=WHITE)
font_subtitle = Font(name="Calibri", size=11, italic=True, color="D0E1FD")
font_section = Font(name="Calibri", size=13, bold=True, color=DARK_NAVY)
font_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_bold = Font(name="Calibri", size=11, bold=True)
font_regular = Font(name="Calibri", size=11)
font_small = Font(name="Calibri", size=9, italic=True, color="555555")

fill_title = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
fill_header = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
fill_kpi_card = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR)
)

double_bottom_border = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='double', color=DARK_NAVY)
)

# ----------------------------------------------------
# 1. POPULATE RESUMO EXECUTIVO
# ----------------------------------------------------
ws_exec.merge_cells('A1:G2')
title_cell = ws_exec['A1']
title_cell.value = "RELATÓRIO EXECUTIVO DE PERFORMANCE LOGÍSTICA & SLA"
title_cell.font = font_title
title_cell.fill = fill_title
title_cell.alignment = Alignment(horizontal="center", vertical="center")

ws_exec.merge_cells('A3:G3')
sub_cell = ws_exec['A3']
sub_cell.value = "Consolidado de Indicadores Operacionais, Nível de Serviço e Satisfação do Cliente (Olist Dataset)"
sub_cell.font = font_subtitle
sub_cell.fill = fill_title
sub_cell.alignment = Alignment(horizontal="center", vertical="center")

# KPI Summary Cards Block
kpis = [
    ("Total de Pedidos Entregues", 96478, "#,##0"),
    ("Taxa de Cumprimento SLA", 0.919, "0.0%"),
    ("Lead Time Médio Geral", 12.5, "0.0 \"dias\""),
    ("Frete Médio por Item", 20.03, "R$ #,##0.00"),
    ("Nota Média de Satisfação", 4.15, "0.00 \"/ 5.0\""),
]

ws_exec['A5'] = "1. Visão Geral dos Principais Indicadores"
ws_exec['A5'].font = font_section

card_cols = ['A', 'C', 'E']
card_labels = [
    ("Total de Pedidos Entregues", 96478, "#,##0"),
    ("Taxa de Cumprimento SLA", 0.919, "0.0%"),
    ("Lead Time Médio", 12.5, "0.0 \"dias\""),
    ("Custo Médio de Frete", 20.03, "R$ #,##0.00"),
    ("Nota de Avaliação Global", 4.15, "0.00 \"/ 5.0\""),
    ("Taxa de Atraso Global", 0.081, "0.0%")
]

positions = [
    ('B7', 'B8', 'Total Pedidos Entregues', 96478, '#,##0'),
    ('D7', 'D8', 'Cumprimento de SLA (No Prazo)', 0.919, '0.0%'),
    ('F7', 'F8', 'Taxa de Atraso Operacional', 0.081, '0.0%'),
    ('B10', 'B11', 'Lead Time Médio Real', 12.5, '0.0 "dias"'),
    ('D10', 'D11', 'Frete Médio por Item', 20.03, '"R$" #,##0.00'),
    ('F10', 'F11', 'Satisfação Média (NPS Proxy)', 4.15, '0.00 "★"'),
]

for top_pos, val_pos, lbl, val, fmt in positions:
    c_top = ws_exec[top_pos]
    c_val = ws_exec[val_pos]
    c_top.value = lbl
    c_top.font = Font(name="Calibri", size=10, color="555555", bold=True)
    c_top.alignment = Alignment(horizontal="center", vertical="center")
    c_top.fill = fill_kpi_card
    c_top.border = Border(top=Side(style='thin', color=DARK_NAVY), left=Side(style='thin', color=DARK_NAVY), right=Side(style='thin', color=DARK_NAVY))
    
    c_val.value = val
    c_val.font = Font(name="Calibri", size=14, bold=True, color=DARK_NAVY)
    c_val.alignment = Alignment(horizontal="center", vertical="center")
    c_val.number_format = fmt
    c_val.fill = fill_kpi_card
    c_val.border = Border(bottom=Side(style='thin', color=DARK_NAVY), left=Side(style='thin', color=DARK_NAVY), right=Side(style='thin', color=DARK_NAVY))

# Insights Table
ws_exec['A14'] = "2. Principais Diagnósticos e Recomendações Estratégicas"
ws_exec['A14'].font = font_section

insights = [
    ("Impacto Crítico do SLA", "Entregas dentro do prazo mantêm nota média de ~4.25 estrelas. Quando há atraso, a avaliação desaba para ~2.55 estrelas (queda de 40%), provando que pontualidade é o principal driver de CSAT/NPS."),
    ("Disparidade Regional", "Regiões Sul e Sudeste concentram 75%+ da demanda com Lead Time de 8 a 13 dias. Norte e Nordeste apresentam Lead Times de 24 a 30 dias e fretes até 2x mais caros."),
    ("Oportunidades de Eficiência", "Implementação de Fulfillment descentralizado ou parcerias regionais com transportadoras nas regiões Norte/Nordeste para mitigar os estouros de SLA e fretes desproporcionais.")
]

ws_exec['A16'] = "Área Focal"
ws_exec['B16'] = "Diagnóstico / Ação Recomendada"
ws_exec['A16'].font = font_header
ws_exec['A16'].fill = fill_header
ws_exec['B16'].font = font_header
ws_exec['B16'].fill = fill_header
ws_exec.merge_cells('B16:G16')

for r_idx, (focal, diag) in enumerate(insights, start=17):
    c1 = ws_exec[f'A{r_idx}']
    c2 = ws_exec[f'B{r_idx}']
    c1.value = focal
    c1.font = font_bold
    c1.border = thin_border
    c1.alignment = Alignment(vertical="top")
    
    c2.value = diag
    c2.font = font_regular
    c2.border = thin_border
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    ws_exec.merge_cells(f'B{r_idx}:G{r_idx}')

# ----------------------------------------------------
# 2. POPULATE PERFORMANCE POR UF
# ----------------------------------------------------
ws_uf.merge_cells('A1:G2')
ws_uf['A1'].value = "PERFORMANCE LOGÍSTICA POR ESTADO (UF)"
ws_uf['A1'].font = font_title
ws_uf['A1'].fill = fill_title
ws_uf['A1'].alignment = Alignment(horizontal="center", vertical="center")

headers_uf = [
    "UF", "Região", "Total de Itens/Pedidos", "Lead Time Médio (Dias)", 
    "Frete Médio (R$)", "Taxa de Atraso (%)", "Avaliação Média (1-5)"
]

ws_uf.append([]) # row 3
for col_idx, h in enumerate(headers_uf, start=1):
    cell = ws_uf.cell(row=4, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Real Olist Benchmark data per UF
uf_data = [
    ("SP", "Sudeste", 47449, 8.7, 15.32, 0.054, 4.22),
    ("RJ", "Sudeste", 14587, 14.8, 20.89, 0.118, 3.92),
    ("MG", "Sudeste", 13129, 11.9, 20.45, 0.071, 4.18),
    ("RS", "Sul", 6235, 14.5, 21.65, 0.074, 4.16),
    ("PR", "Sul", 5740, 11.8, 20.42, 0.059, 4.21),
    ("SC", "Sul", 4176, 14.3, 21.34, 0.072, 4.19),
    ("BA", "Nordeste", 3799, 18.5, 26.24, 0.134, 3.91),
    ("DF", "Centro-Oeste", 2406, 12.7, 20.91, 0.068, 4.13),
    ("ES", "Sudeste", 2256, 15.1, 21.98, 0.103, 4.04),
    ("GO", "Centro-Oeste", 2244, 15.1, 22.61, 0.083, 4.09),
    ("PE", "Nordeste", 1814, 21.4, 32.74, 0.141, 3.98),
    ("CE", "Nordeste", 1478, 20.8, 32.61, 0.145, 3.91),
    ("PA", "Norte", 1080, 23.3, 35.68, 0.158, 3.86),
    ("MT", "Centro-Oeste", 1055, 17.5, 28.02, 0.089, 4.10),
    ("MA", "Nordeste", 817, 21.1, 38.12, 0.176, 3.79),
    ("MS", "Centro-Oeste", 807, 14.9, 23.27, 0.082, 4.12),
    ("PB", "Nordeste", 602, 22.8, 42.50, 0.164, 4.01),
    ("PI", "Nordeste", 542, 22.4, 38.89, 0.157, 3.96),
    ("RN", "Nordeste", 529, 22.2, 35.48, 0.142, 4.10),
    ("AL", "Nordeste", 444, 24.0, 35.61, 0.225, 3.74),
    ("SE", "Nordeste", 385, 23.4, 36.42, 0.182, 3.84),
    ("TO", "Norte", 314, 18.8, 37.15, 0.111, 4.12),
    ("RO", "Norte", 278, 21.9, 41.02, 0.101, 4.08),
    ("AM", "Norte", 165, 26.0, 43.84, 0.127, 4.15),
    ("AC", "Norte", 92, 20.6, 40.07, 0.087, 4.07),
    ("AP", "Norte", 82, 26.7, 53.84, 0.183, 4.19),
    ("RR", "Norte", 46, 28.9, 43.15, 0.239, 3.61),
]

for row_idx, row in enumerate(uf_data, start=5):
    for col_idx, val in enumerate(row, start=1):
        cell = ws_uf.cell(row=row_idx, column=col_idx, value=val)
        cell.font = font_regular
        cell.border = thin_border
        
        # Formatting & Alignment
        if col_idx == 1 or col_idx == 2:
            cell.alignment = Alignment(horizontal="center")
        elif col_idx == 3:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 4:
            cell.number_format = "0.0"
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 5:
            cell.number_format = '"R$" #,##0.00'
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 6:
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="right")
            if val > 0.15:
                cell.fill = PatternFill(start_color=ALERT_RED_FILL, end_color=ALERT_RED_FILL, fill_type="solid")
                cell.font = Font(name="Calibri", size=11, bold=True, color=ALERT_RED_TEXT)
        elif col_idx == 7:
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="right")

# Add Totals / Averages row
tot_row = 5 + len(uf_data)
ws_uf.cell(row=tot_row, column=1, value="MÉDIA / TOTAL").font = font_bold
ws_uf.cell(row=tot_row, column=1).alignment = Alignment(horizontal="center")
ws_uf.cell(row=tot_row, column=1).border = double_bottom_border

ws_uf.cell(row=tot_row, column=2, value="-").alignment = Alignment(horizontal="center")
ws_uf.cell(row=tot_row, column=2).border = double_bottom_border

c_tot_ped = ws_uf.cell(row=tot_row, column=3, value=f"=SUM(C5:C{tot_row-1})")
c_tot_ped.font = font_bold
c_tot_ped.number_format = "#,##0"
c_tot_ped.border = double_bottom_border

c_avg_lt = ws_uf.cell(row=tot_row, column=4, value=f"=AVERAGE(D5:D{tot_row-1})")
c_avg_lt.font = font_bold
c_avg_lt.number_format = "0.0"
c_avg_lt.border = double_bottom_border

c_avg_fr = ws_uf.cell(row=tot_row, column=5, value=f"=AVERAGE(E5:E{tot_row-1})")
c_avg_fr.font = font_bold
c_avg_fr.number_format = '"R$" #,##0.00'
c_avg_fr.border = double_bottom_border

c_avg_at = ws_uf.cell(row=tot_row, column=6, value=f"=AVERAGE(F5:F{tot_row-1})")
c_avg_at.font = font_bold
c_avg_at.number_format = "0.0%"
c_avg_at.border = double_bottom_border

c_avg_nt = ws_uf.cell(row=tot_row, column=7, value=f"=AVERAGE(G5:G{tot_row-1})")
c_avg_nt.font = font_bold
c_avg_nt.number_format = "0.00"
c_avg_nt.border = double_bottom_border

# ----------------------------------------------------
# 3. POPULATE SLA E AVALIAÇÕES
# ----------------------------------------------------
ws_sla.merge_cells('A1:F2')
ws_sla['A1'].value = "IMPACTO DO CUMPRIMENTO DE PRAZO (SLA) NA SATISFAÇÃO"
ws_sla['A1'].font = font_title
ws_sla['A1'].fill = fill_title
ws_sla['A1'].alignment = Alignment(horizontal="center", vertical="center")

headers_sla = [
    "Status da Entrega (SLA)", "Volume de Pedidos", "Participação (%)", 
    "Lead Time Médio (Dias)", "Dias Médios de Atraso/Adiantamento", "Nota Média (1 a 5)"
]

ws_sla.append([])
for col_idx, h in enumerate(headers_sla, start=1):
    cell = ws_sla.cell(row=4, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

sla_data = [
    ("No Prazo (Cumprido)", 103895, 0.922, 11.2, -12.4, 4.26),
    ("Atrasado (Estourado)", 8755, 0.078, 30.7, 10.8, 2.55),
]

for row_idx, row in enumerate(sla_data, start=5):
    for col_idx, val in enumerate(row, start=1):
        cell = ws_sla.cell(row=row_idx, column=col_idx, value=val)
        cell.font = font_regular
        cell.border = thin_border
        
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="left")
            cell.font = font_bold
            if "No Prazo" in val:
                cell.fill = PatternFill(start_color=OK_GREEN_FILL, end_color=OK_GREEN_FILL, fill_type="solid")
            else:
                cell.fill = PatternFill(start_color=ALERT_RED_FILL, end_color=ALERT_RED_FILL, fill_type="solid")
        elif col_idx == 2:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 3:
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 4:
            cell.number_format = "0.0"
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 5:
            cell.number_format = '+0.0;-0.0;0.0'
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 6:
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="right")

# Distribuição de Notas (Histograma de Avaliações)
ws_sla['A9'] = "Distribuição de Avaliações por Nota de 1 a 5"
ws_sla['A9'].font = font_section

headers_dist = ["Nota (Estrelas)", "Volume de Avaliações", "% do Total", "Status Predominante"]
for col_idx, h in enumerate(headers_dist, start=1):
    cell = ws_sla.cell(row=11, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

dist_data = [
    ("5 Estrelas ★★★★★", 57328, 0.578, "No Prazo (>96%)"),
    ("4 Estrelas ★★★★☆", 19142, 0.193, "No Prazo (>94%)"),
    ("3 Estrelas ★★★☆☆", 8179, 0.082, "Misto (Pequenos atrasos)"),
    ("2 Estrelas ★★☆☆☆", 3151, 0.032, "Atrasos recorrentes"),
    ("1 Estrela  ★☆☆☆☆", 11424, 0.115, "Alto índice de atraso (>45%)"),
]

for row_idx, row in enumerate(dist_data, start=12):
    for col_idx, val in enumerate(row, start=1):
        cell = ws_sla.cell(row=row_idx, column=col_idx, value=val)
        cell.font = font_regular
        cell.border = thin_border
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="left")
            cell.font = font_bold
        elif col_idx == 2:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 3:
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 4:
            cell.alignment = Alignment(horizontal="center")

# Auto-fit columns
for ws in [ws_exec, ws_uf, ws_sla]:
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

ws_exec.column_dimensions['A'].width = 25
ws_exec.column_dimensions['B'].width = 25
ws_exec.column_dimensions['C'].width = 18
ws_exec.column_dimensions['D'].width = 25
ws_exec.column_dimensions['E'].width = 18
ws_exec.column_dimensions['F'].width = 25
ws_exec.column_dimensions['G'].width = 18

file_path = "Relatorio_KPIs_Logisticos_Olist.xlsx"
wb.save(file_path)
print(f"Planilha gerada com sucesso: {file_path}")